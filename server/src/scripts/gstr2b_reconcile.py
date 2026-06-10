#!/usr/bin/env python3
"""GSTR-2B Reconciliation Script.
Reads JSON from stdin, writes JSON to stdout.
Phase 'parse' : parse files, run deterministic matching, return results + tier2.5 candidates.
Phase 'generate': apply Claude results, produce 5-sheet Excel, return summary + entries.
"""
import sys, json, re, os
from datetime import datetime, date

try:
    import openpyxl
    from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
    from openpyxl.utils import get_column_letter
except ImportError:
    json.dump({"error": "openpyxl not installed. Run: pip3 install openpyxl"}, sys.stdout)
    sys.exit(1)

# ── Colour constants (no leading #) ──────────────────────────────────────────
NAVY         = "1F3864"
NAVY2        = "2E4B8A"
WHITE_HEX    = "FFFFFF"
GREEN_ROW    = "E2EFDA"
YELLOW_ROW   = "FFEB9C"
BLUE_ROW     = "BDD7EE"
PURPLE_ROW   = "E2E2F0"
ALT_GREEN    = "F0FDF4"
ALT_AMBER    = "FFFBEB"
RED_TOTAL    = "C00000"
RED_DIFF     = "FFC7CE"

def _fill(h):  return PatternFill(start_color=h, end_color=h, fill_type="solid")
def _font(bold=False, color=None, size=10):
    return Font(name="Arial", bold=bold, color=(color or "000000"), size=size)
def _border():
    s = Side(style="thin", color="D3D3D3")
    return Border(left=s, right=s, top=s, bottom=s)

# ── Name normalisation ────────────────────────────────────────────────────────
def norm(s):
    s = str(s).upper()
    s = re.sub(r'[^A-Z0-9 ]', ' ', s)
    s = re.sub(r'\bAND\b', ' ', s)   # treat "AND" same as "&" (both stripped)
    return re.sub(r'\s+', ' ', s).strip()

def token_overlap(a, b):
    ta = {w for w in norm(a).split() if len(w) >= 3}
    tb = {w for w in norm(b).split() if len(w) >= 3}
    if not ta and not tb: return 1.0
    if not ta or not tb:  return 0.0
    return len(ta & tb) / max(len(ta), len(tb))

# Generic trade words shared by unrelated firms ("X AGENCIES" vs "Y AGENCIES").
# Distinctive-token overlap prevents these from creating phantom matches.
GENERIC_TOKENS = {
    'AGENCIES','AGENCY','TRADERS','TRADING','STORES','STORE','ENTERPRISES',
    'ENTERPRISE','ASSOCIATES','ASSOCIATE','INDUSTRIES','DISTRIBUTORS',
    'DISTRIBUTION','MARKETING','FOODS','FOOD','PRODUCTS','MILLS','MILL',
    'SNACKS','CORPORATION','VENTURES','COMPANY','CO','LIMITED','LTD','PRIVATE',
    'PVT','LLP','SONS','TVL','CASH','THE','SRI','SHRI','SREE','SHREE',
}

def distinct_overlap(a, b):
    """Overlap coefficient on distinctive (non-generic) name tokens."""
    da = {w for w in norm(a).split() if w not in GENERIC_TOKENS}
    db = {w for w in norm(b).split() if w not in GENERIC_TOKENS}
    if not da or not db: return 0.0
    return len(da & db) / min(len(da), len(db))

def name_score(a, b):
    """Combined name similarity: distinctive tokens dominate, generic fallback."""
    d = distinct_overlap(a, b)
    return d if d > 0 else token_overlap(a, b) * 0.5

# ── Bill-number helpers ───────────────────────────────────────────────────────
BILL_NO_RE = re.compile(
    r'(BILL|CREDIT\s*NOTE|DEBIT\s*NOTE|INV(?:OICE)?)\s*NO\.?\s*([A-Za-z0-9/\-\.]+)',
    re.IGNORECASE)

def extract_bill_no(narration):
    m = BILL_NO_RE.search(narration or '')
    return m.group(2).strip().rstrip('.') if m else ''

def doc_type_of(narration):
    u = (narration or '').upper()
    if 'CREDIT NOTE' in u: return 'Credit Note'
    if 'DEBIT NOTE'  in u: return 'Debit Note'
    return 'Invoice'

def _digits(s): return re.sub(r'\D', '', s or '')

def bill_no_match(inv_no, books_bill):
    """True when the books bill number is the same document as the 2B invoice no.
    GSTR-1 numbers often wrap the plain Tally bill no in prefixes/FY suffixes:
    books '1058' ↔ 2B '1058'; books '142' ↔ 2B '26-27/00142'; books '63' ↔
    2B 'TUP01/V2000063'."""
    d = _digits(books_bill).lstrip('0')
    if not d: return False
    e = _digits(inv_no)
    if e.lstrip('0') == d: return True
    if len(d) >= 2 and e.endswith(d): return True
    return d in [g.lstrip('0') for g in re.findall(r'\d+', inv_no or '')]

NARR_NOISE = ["AS PER GSTR 2B","AS PER GSTR2B","BILL NO","BILL NO.",
              "INV NO","INVOICE NO","VIDE","BEING","TOWARDS",
              "PAYMENT TO","RECEIVED FROM"]

def clean_narration(s):
    s = norm(s)
    for n in NARR_NOISE: s = s.replace(n, ' ')
    return re.sub(r'\s+', ' ', s).strip()

# ── Utility helpers ───────────────────────────────────────────────────────────
def _flt(v):
    try:    return round(float(v or 0), 2)
    except: return 0.0

def _str(v): return str(v or '').strip()

def to_date_str(v):
    if v is None: return ''
    if isinstance(v, (datetime, date)): return v.strftime('%d/%m/%Y')
    return _str(v)

def get_rate_label(inv):
    taxable = inv.get('taxable_value', 0) or 0
    cgst    = inv.get('cgst', 0) or 0
    if taxable > 0 and cgst > 0:
        r = round(cgst / taxable * 100, 1)
        if r <= 1.6: return '1.5%'
        if r <= 2.6: return '2.5%'
    return '9%'

# ── GSTR-2B parser ────────────────────────────────────────────────────────────
def parse_gstr2b(path):
    wb = openpyxl.load_workbook(path, data_only=True)

    company_info = {}
    if 'Read me' in wb.sheetnames:
        r = list(wb['Read me'].iter_rows(values_only=True))
        def _c(ri, ci): return _str(r[ri][ci]) if len(r) > ri and len(r[ri]) > ci else ''
        company_info = {
            'fy':         _c(3, 2),
            'period':     _c(4, 2),
            'gstin':      _c(5, 2),
            'legal_name': _c(6, 2),
            'trade_name': _c(7, 2),
        }

    if 'B2B' not in wb.sheetnames:
        raise ValueError("GSTR-2B file has no 'B2B' sheet")

    rows = list(wb['B2B'].iter_rows(values_only=True))
    if len(rows) < 7:
        raise ValueError("B2B sheet has too few rows")

    def _mk(gstin, name, doc_no, doc_date, pos, taxable, igst, cgst, sgst, itc, doc_type):
        inv = {
            'gstin':           gstin,
            'supplier_name':   name,
            'invoice_no':      doc_no,
            'invoice_date':    doc_date,
            'taxable_value':   taxable,
            'igst':            igst,
            'cgst':            cgst,
            'sgst':            sgst,
            'place_of_supply': pos,
            'itc_availability':itc,
            'doc_type':        doc_type,
        }
        # Classify by the tax actually charged, not by GSTIN prefix — a local
        # supplier can still bill IGST (and vice versa).
        if   cgst > 0: inv['tax_type'] = 'CGST'
        elif igst > 0: inv['tax_type'] = 'IGST'
        else:          inv['tax_type'] = 'CGST' if gstin.startswith('33') else 'IGST'
        inv['tax_amount'] = inv['cgst'] if inv['tax_type'] == 'CGST' else inv['igst']
        inv['rate_label'] = get_rate_label(inv)
        return inv

    # B2B sheet — invoices.
    # Col: 0=GSTIN 1=Name 2=InvNo 3=InvDate 4=InvValue 5=PlaceOfSupply
    #      6=RCM 7=Taxable 8=IGST 9=CGST 10=SGST 11=Cess
    #      12=GSTR1Period 13=GSTR1FilingDate 14=ITCAvailability 15=Reason
    invoices = []
    for row in rows[6:]:
        if not row or len(row) < 11: continue
        gstin = _str(row[0])
        if not gstin or len(gstin) < 10: continue
        invoices.append(_mk(
            gstin, _str(row[1]), _str(row[2]), to_date_str(row[3]), _str(row[5]),
            _flt(row[7]), _flt(row[8]), _flt(row[9]), _flt(row[10]),
            _str(row[14]) if len(row) > 14 else '', 'Invoice'))

    # B2B-CDNR sheet — credit/debit notes (previously ignored entirely).
    # Col: 0=GSTIN 1=Name 2=NoteNo 3=NoteType 4=NoteSupplyType 5=NoteDate
    #      6=NoteValue 7=PlaceOfSupply 8=RCM 9=Taxable 10=IGST 11=CGST 12=SGST 13=Cess
    if 'B2B-CDNR' in wb.sheetnames:
        for row in list(wb['B2B-CDNR'].iter_rows(values_only=True))[6:]:
            if not row or len(row) < 13: continue
            gstin = _str(row[0])
            if not gstin or len(gstin) < 10: continue
            ntype = 'Credit Note' if _str(row[3]).upper().startswith('C') else 'Debit Note'
            invoices.append(_mk(
                gstin, _str(row[1]), _str(row[2]), to_date_str(row[5]), _str(row[7]),
                _flt(row[9]), _flt(row[10]), _flt(row[11]), _flt(row[12]),
                '', ntype))

    return company_info, invoices

# ── Tally ledger parser ───────────────────────────────────────────────────────
def parse_tally_ledger(path):
    wb = openpyxl.load_workbook(path, data_only=True)
    ws = wb.active
    rows = list(ws.iter_rows(values_only=True))
    if len(rows) < 9: return []

    # Find header row: first row that has "DATE" in col 0 AND any CGST/IGST/SGST keyword.
    header_row_idx = 7
    for idx in range(min(len(rows), 20)):
        rv = [_str(v).upper() for v in (rows[idx] or [])]
        has_date = any(v.strip() == 'DATE' for v in rv)
        has_tax  = any(any(kw in v for kw in ['CGST','IGST','SGST']) for v in rv)
        if has_date and has_tax:
            header_row_idx = idx
            break

    headers = [_str(h) for h in (rows[header_row_idx] if len(rows) > header_row_idx else [])]

    # Collect ALL Input CGST and Input IGST column indices.
    # Tally columnar exports show every tax sub-ledger as a separate column on the
    # same row — a single invoice can have both CGST@9% and CGST@2.5% amounts in
    # different columns.  Summing them gives the total CGST per voucher, which is
    # what GSTR-2B reports as a single combined figure.
    cgst_cols, igst_cols = [], []
    for i, h in enumerate(headers):
        if i <= 2: continue
        hu = h.upper()
        if 'CGST' in hu and any(kw in hu for kw in ['INPUT', '@', '-']):
            cgst_cols.append(i)
        elif 'IGST' in hu and any(kw in hu for kw in ['INPUT', '@', '-']):
            igst_cols.append(i)

    # Fallback: any column beyond col 2 that mentions CGST or IGST
    if not cgst_cols and not igst_cols:
        for i, h in enumerate(headers):
            if i <= 2: continue
            if 'CGST' in h.upper(): cgst_cols.append(i)
            elif 'IGST' in h.upper(): igst_cols.append(i)

    default_tax_type = 'IGST' if (igst_cols and not cgst_cols) else 'CGST'

    entries = []
    for row in rows[header_row_idx + 1:]:
        if not row: continue
        col1 = _str(row[1] if len(row) > 1 else '')
        if 'grand total' in col1.lower(): continue
        if not any(row): continue

        party     = _str(row[1] if len(row) > 1 else '')
        narration = _str(row[2] if len(row) > 2 else '')
        date_val  = to_date_str(row[0] if row else None)
        gross     = abs(_flt(row[3] if len(row) > 3 else 0))  # col 3 = Gross Total

        # Sum ALL CGST columns → total CGST for this voucher (matches GSTR-2B combined figure)
        cgst_total = sum(abs(_flt(row[i])) for i in cgst_cols if i < len(row))
        igst_total = sum(abs(_flt(row[i])) for i in igst_cols if i < len(row))

        if cgst_total > 0:
            tax_amt  = cgst_total
            tax_type = 'CGST'
        elif igst_total > 0:
            tax_amt  = igst_total
            tax_type = 'IGST'
        else:
            tax_amt  = 0.0
            tax_type = default_tax_type

        if tax_amt == 0.0 and gross == 0.0: continue
        if not date_val and not party and not narration: continue

        entries.append({
            'date':        date_val,
            'party':       party,
            'narration':   narration,
            'bill_no':     extract_bill_no(narration),
            'doc_type':    doc_type_of(narration),
            'gross_total': gross,
            'tax_amount':  tax_amt,
            'tax_rate':    '',        # not meaningful when summing multiple rate columns
            'tax_type':    tax_type,
            'ledger_file': os.path.basename(path),
            '_used':       False,
        })
    return entries

# ── Matching engine ───────────────────────────────────────────────────────────
# Global tiered passes (strongest evidence first) instead of per-invoice greedy
# matching.  Every tier requires the same doc type (invoice ↔ invoice, credit
# note ↔ credit note) and tax type.  Amount tolerance for a confirmed match is
# ₹1 — the old blanket max(5%, ₹50) tolerance created phantom matches between
# unrelated suppliers sharing generic words like "AGENCIES".
def _eligible(inv, e):
    return (not e['_used']
            and e['tax_type'] == inv['tax_type']
            and e['doc_type'] == inv.get('doc_type', 'Invoice')
            and e['tax_amount'] > 0)

def run_match_passes(invoices, tally_entries):
    """Returns {invoice index → (status, match_type, tally index, reason, is_t25)}."""
    out = {}

    def tier(label, cond, status='matched', unique_only=False):
        for ii, inv in enumerate(invoices):
            if ii in out or inv['tax_amount'] <= 0: continue
            cands = [(ti, e) for ti, e in enumerate(tally_entries)
                     if _eligible(inv, e) and cond(inv, e)]
            if not cands: continue
            if unique_only and len(cands) > 1: continue
            ti, e = max(cands, key=lambda c: name_score(inv['supplier_name'], c[1]['party']))
            tally_entries[ti]['_used'] = True
            out[ii] = (status, label, ti, '', False)

    # T0 — GSTIN written in the Tally narration/party + amount agrees
    tier('GSTIN Match',
         lambda inv, e: inv['gstin'] in (e['narration'] + ' ' + e['party']).upper()
                        and abs(e['tax_amount'] - inv['tax_amount']) <= 1.0)

    # T1 — name + bill number + amount: the strongest everyday signal
    tier('Exact (name+bill+amount)',
         lambda inv, e: name_score(inv['supplier_name'], e['party']) >= 0.5
                        and bill_no_match(inv['invoice_no'], e['bill_no'])
                        and abs(e['tax_amount'] - inv['tax_amount']) <= 1.0)

    # T2 — name + amount (unique candidate; bill format may differ entirely)
    tier('Name + Amount',
         lambda inv, e: name_score(inv['supplier_name'], e['party']) >= 0.5
                        and abs(e['tax_amount'] - inv['tax_amount']) <= 1.0,
         unique_only=True)

    # T3 — bill number + amount (unique): covers cash-ledger entries where the
    # Tally party is blank or just "Cash" but the narration carries the bill no
    tier('Bill + Amount',
         lambda inv, e: bill_no_match(inv['invoice_no'], e['bill_no'])
                        and abs(e['tax_amount'] - inv['tax_amount']) <= 1.0,
         unique_only=True)

    # T4 — narration carries the supplier name (party blank/cash) + amount
    tier('Via Narration',
         lambda inv, e: name_score(inv['supplier_name'], e['party']) < 0.5
                        and token_overlap(inv['supplier_name'], clean_narration(e['narration'])) >= 0.4
                        and abs(e['tax_amount'] - inv['tax_amount']) <= 1.0,
         unique_only=True)

    # T5 — consolidated bookings: several 2B docs from one supplier summing to
    # a single unused Tally voucher (e.g. 9 small invoices booked as one entry)
    from itertools import combinations
    for ti, e in enumerate(tally_entries):
        if e['_used'] or e['tax_amount'] <= 0: continue
        pool = [ii for ii, inv in enumerate(invoices)
                if ii not in out and inv['tax_amount'] > 0 and _eligible(inv, e)
                and distinct_overlap(inv['supplier_name'], e['party']) >= 0.99]
        if not (2 <= len(pool) <= 14): continue
        hit = None
        for size in range(len(pool), 1, -1):
            for combo in combinations(pool, size):
                if abs(sum(invoices[ii]['tax_amount'] for ii in combo) - e['tax_amount']) <= 1.0:
                    hit = combo; break
            if hit: break
        if hit:
            e['_used'] = True
            for ii in hit:
                out[ii] = ('matched', f'Consolidated ({len(hit)} docs = 1 voucher)', ti,
                           f'Voucher {e["bill_no"] or e["narration"]} covers {len(hit)} GSTR-2B documents', False)

    # T6 — same supplier + same bill number but amount differs → review
    for ii, inv in enumerate(invoices):
        if ii in out or inv['tax_amount'] <= 0: continue
        cands = [(ti, e) for ti, e in enumerate(tally_entries)
                 if _eligible(inv, e)
                 and name_score(inv['supplier_name'], e['party']) >= 0.5
                 and bill_no_match(inv['invoice_no'], e['bill_no'])]
        if cands:
            ti, e = cands[0]
            tally_entries[ti]['_used'] = True
            diff = round(abs(e['tax_amount'] - inv['tax_amount']), 2)
            out[ii] = ('unmatched', 'Amount Mismatch', ti,
                       f'Same supplier & bill no, tax differs by ₹{diff:,.2f} — verify invoice', False)

    # T7 — offsetting invoice/credit-note pairs in 2B with no books entry:
    # supplier billed and fully reversed; net zero, booking optional
    for ii, inv in enumerate(invoices):
        if ii in out or inv['tax_amount'] <= 0 or inv.get('doc_type') != 'Invoice': continue
        for jj, cn in enumerate(invoices):
            if jj in out or cn.get('doc_type') != 'Credit Note': continue
            if (cn['gstin'] == inv['gstin'] and cn['tax_type'] == inv['tax_type']
                    and abs(cn['tax_amount'] - inv['tax_amount']) <= 0.02):
                out[ii] = ('not_in_books', 'Offset Pair', None,
                           f'Cancelled by credit note {cn["invoice_no"]} — net zero, booking optional', False)
                out[jj] = ('not_in_books', 'Offset Pair', None,
                           f'Cancels invoice {inv["invoice_no"]} — net zero, booking optional', False)
                break

    # T8 — ambiguous: plausible supplier-name candidate within 10% → Claude
    for ii, inv in enumerate(invoices):
        if ii in out or inv['tax_amount'] <= 0: continue
        cands = []
        for ti, e in enumerate(tally_entries):
            if not _eligible(inv, e): continue
            ds  = distinct_overlap(inv['supplier_name'], e['party'])
            nsc = token_overlap(inv['supplier_name'], clean_narration(e['narration']))
            diff = abs(e['tax_amount'] - inv['tax_amount'])
            if (ds >= 0.5 or nsc >= 0.4) and diff <= max(inv['tax_amount'] * 0.10, 5):
                cands.append((ti, e, ds, nsc, diff))
        if cands:
            ti, e, ds, nsc, diff = max(cands, key=lambda c: c[2] + c[3])
            out[ii] = ('tier2_5', 'Pending AI Review', ti,
                       f'Ambiguous: name={ds:.2f}, narr={nsc:.2f}, diff=₹{diff:,.2f}', True)

    # Everything else: zero-tax docs are informational; the rest are not booked
    for ii, inv in enumerate(invoices):
        if ii in out: continue
        if inv['tax_amount'] <= 0:
            out[ii] = ('not_in_books', 'Nil Tax', None,
                       'Nil tax in GSTR-2B — no ITC impact', False)
        else:
            out[ii] = ('not_in_books', 'No Match', None,
                       'Not found in purchase register', False)
    return out

# ── Phase 1: parse + deterministic match ─────────────────────────────────────
def run_parse(gstr2b_path, ledger_paths):
    company_info, gstr2b_invoices = parse_gstr2b(gstr2b_path)

    all_tally = []
    for lp in ledger_paths:
        all_tally.extend(parse_tally_ledger(lp))

    # Deduplicate across files: a voucher that posts to two tax-rate ledgers
    # (e.g. an invoice with both 5% and 18% lines) appears as a full identical
    # row in BOTH register exports.  Narration is part of the key so two
    # different bills from the same party/date/amount are NOT collapsed.
    seen_keys: set = set()
    deduped = []
    for e in all_tally:
        key = (e['date'], norm(e['party']), norm(e['narration']),
               e['gross_total'], e['tax_type'])
        if key not in seen_keys:
            seen_keys.add(key)
            deduped.append(e)
    all_tally = deduped

    match_map = run_match_passes(gstr2b_invoices, all_tally)

    results, tier2_5_cands = [], []
    for inv_idx, inv in enumerate(gstr2b_invoices):
        status, mtype, tidx, reason, is_t25 = match_map[inv_idx]

        me = all_tally[tidx] if tidx is not None else None

        proposed = None
        if is_t25 and me:
            proposed = {
                'party':      me['party'],
                'narration':  me['narration'],
                'date':       me['date'],
                'tax_amount': me['tax_amount'],
                'ledger_file':me['ledger_file'],
            }

        result = {
            **{k: v for k, v in inv.items()},
            'status':           status,
            'match_type':       mtype,
            'books_party':      me['party']       if me else None,
            'books_date':       me['date']        if me else None,
            'books_ledger':     me['ledger_file'] if me else None,
            'books_tax':        me['tax_amount']  if me else None,
            'books_narration':  me['narration']   if me else None,
            'tax_diff':         0.0 if mtype.startswith('Consolidated')
                                else round(abs((me['tax_amount'] if me else 0) - inv['tax_amount']), 2),
            'discrepancy_reason': reason,
            'tier2_5':          is_t25,
            '_proposed_tally':  proposed,
        }
        results.append(result)

        if is_t25 and me:
            tier2_5_cands.append({
                'inv_no': inv['invoice_no'],
                'gstr2b': {
                    'gstin':          inv['gstin'],
                    'supplier_name':  inv['supplier_name'],
                    'invoice_no':     inv['invoice_no'],
                    'invoice_date':   inv['invoice_date'],
                    'taxable_value':  inv['taxable_value'],
                    'tax_amount':     inv['tax_amount'],
                    'tax_type':       inv['tax_type'],
                    'place_of_supply':inv['place_of_supply'],
                },
                'tally_candidates': [{
                    'party':     me['party'],
                    'narration': me['narration'],
                    'date':      me['date'],
                    'tax_amount':me['tax_amount'],
                    'ledger':    me['ledger_file'],
                }],
                'reason_for_review': reason,
            })

    # books_only = unused Tally entries, excluding tier2.5 proposed candidates
    proposed_keys = {
        (r['_proposed_tally']['party'],
         r['_proposed_tally']['ledger_file'],
         r['_proposed_tally']['tax_amount'],
         r['_proposed_tally']['date'])
        for r in results if r.get('_proposed_tally')
    }
    books_only = [
        {'party': e['party'], 'narration': e['narration'], 'date': e['date'],
         'tax_amount': e['tax_amount'], 'tax_type': e['tax_type'], 'ledger': e['ledger_file']}
        for e in all_tally
        if not e['_used'] and e['tax_amount'] > 0
        and (e['party'], e['ledger_file'], e['tax_amount'], e['date']) not in proposed_keys
    ]

    return {
        'company_info':       company_info,
        'results':            results,
        'tier2_5_candidates': tier2_5_cands,
        'books_only':         books_only,
    }

# ── Apply Claude results ──────────────────────────────────────────────────────
def apply_claude_results(parse_result, claude_results):
    cmap     = {r['inv_no']: r for r in (claude_results or [])}
    results  = parse_result['results']
    books_only = list(parse_result.get('books_only', []))

    for r in results:
        if not r.get('tier2_5'): continue

        proposed = r.get('_proposed_tally') or {}
        cr       = cmap.get(r['invoice_no'])

        if cr and not cr.get('fallback'):
            conf     = int(cr.get('confidence', 0) or 0)
            is_match = bool(cr.get('match', False))

            if is_match and conf >= 8:
                r['status']    = 'matched'
                r['match_type'] = 'AI-Assisted'
                r['books_party']  = proposed.get('party')
                r['books_date']   = proposed.get('date')
                r['books_ledger'] = proposed.get('ledger_file')
                r['books_tax']    = proposed.get('tax_amount')
                r['tax_diff']     = round(abs((proposed.get('tax_amount') or 0) - r['tax_amount']), 2)
                r['discrepancy_reason'] = cr.get('reason', '')
            elif is_match and conf >= 5:
                r['status']    = 'unmatched'
                r['match_type'] = 'AI-Assisted'
                r['discrepancy_reason'] = (cr.get('reason', '') or '') + ' (confidence 5–7: manual review needed)'
            else:
                r['status']    = 'not_in_books'
                r['match_type'] = 'AI-Assisted'
                r['discrepancy_reason'] = cr.get('reason', '')
                if proposed:
                    books_only.append({
                        'party':     proposed.get('party', ''),
                        'narration': proposed.get('narration', ''),
                        'date':      proposed.get('date', ''),
                        'tax_amount':proposed.get('tax_amount', 0),
                        'tax_type':  r.get('tax_type', 'CGST'),
                        'ledger':    proposed.get('ledger_file', ''),
                    })
        else:
            # Fallback: Claude unavailable
            r['match_type'] = 'Fallback — AI unavailable'
            t_tax = proposed.get('tax_amount', 0) or 0
            if abs(t_tax - r['tax_amount']) <= 10:
                r['status'] = 'unmatched'
                r['discrepancy_reason'] = 'AI unavailable — amount-only match, manual review required'
            else:
                r['status'] = 'not_in_books'
                if proposed:
                    books_only.append({
                        'party':     proposed.get('party', ''),
                        'narration': proposed.get('narration', ''),
                        'date':      proposed.get('date', ''),
                        'tax_amount':proposed.get('tax_amount', 0),
                        'tax_type':  r.get('tax_type', 'CGST'),
                        'ledger':    proposed.get('ledger_file', ''),
                    })

    # Strip internal fields
    for r in results:
        r.pop('_proposed_tally', None)
        r.pop('tier2_5', None)

    parse_result['books_only'] = books_only
    return parse_result

# ── Excel: shared helpers ─────────────────────────────────────────────────────
def _title_row(ws, row, text, ncols, height=24, size=12):
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=ncols)
    c = ws.cell(row=row, column=1, value=text)
    c.fill = _fill(NAVY); c.font = _font(bold=True, color=WHITE_HEX, size=size)
    c.alignment = Alignment(horizontal='center', vertical='center')
    ws.row_dimensions[row].height = height

def _header_row(ws, row, headers):
    for ci, h in enumerate(headers, 1):
        c = ws.cell(row=row, column=ci, value=h)
        c.fill = _fill(NAVY); c.font = _font(bold=True, color=WHITE_HEX)
        c.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        c.border = _border()
    ws.row_dimensions[row].height = 22

def _data_row(ws, row, vals, bg=None, amt_cols=None):
    for ci, v in enumerate(vals, 1):
        c = ws.cell(row=row, column=ci, value=v)
        if bg: c.fill = _fill(bg)
        c.font = _font(); c.alignment = Alignment(vertical='center')
        c.border = _border()
        if amt_cols and ci in amt_cols:
            c.number_format = '#,##0.00'
    ws.row_dimensions[row].height = 16

def _autofit(ws, min_w=12, max_w=52):
    for col in ws.columns:
        mx = max((len(str(c.value)) for c in col if c.value is not None), default=0)
        ws.column_dimensions[get_column_letter(col[0].column)].width = min(max_w, max(min_w, mx + 2))

def _freeze(ws, row=3, col=1):
    ws.freeze_panes = ws.cell(row=row, column=col)

def _autofilter(ws, header_row, ncols):
    ws.auto_filter.ref = f"A{header_row}:{get_column_letter(ncols)}{header_row}"

# ── Sheet 1: Summary ──────────────────────────────────────────────────────────
def build_summary_sheet(wb, company, mc, uc, nc, bc, mi, ui, ni, bi):
    ws  = wb.create_sheet("Summary")
    co  = company.get('legal_name') or company.get('trade_name') or 'Company'
    per = company.get('period', '')
    fy  = company.get('fy', '')

    _title_row(ws, 1, f"{co} — GSTR-2B Reconciliation Report", 6, height=26, size=13)

    sub = f"GSTIN: {company.get('gstin','')}   |   Period: {per}   |   FY: {fy}"
    ws.merge_cells('A2:F2')
    c2 = ws.cell(row=2, column=1, value=sub)
    c2.fill = _fill(NAVY2); c2.font = _font(bold=False, color=WHITE_HEX, size=9)
    c2.alignment = Alignment(horizontal='center', vertical='center')
    ws.row_dimensions[2].height = 15

    ws.row_dimensions[3].height = 8

    def _proof_table(start_row, label, matched, unmatched, not_books, books):
        # Section header
        ws.merge_cells(start_row=start_row, start_column=1, end_row=start_row, end_column=6)
        ws.cell(row=start_row, column=1, value=label).fill = _fill(NAVY)
        ws.cell(row=start_row, column=1).font = _font(bold=True, color=WHITE_HEX, size=11)
        ws.cell(row=start_row, column=1).alignment = Alignment(horizontal='left', vertical='center')
        ws.row_dimensions[start_row].height = 22

        hdrs = ['Category', 'Count', 'GSTR-2B Taxable ₹', 'GSTR-2B Tax ₹', 'Books Tax ₹', 'Difference ₹']
        _header_row(ws, start_row + 1, hdrs)

        def _row(rn, cat, objs, books_list, bg, is_books=False):
            cnt  = len(objs) if not is_books else len(books_list)
            g_tx = round(sum(r['taxable_value'] for r in objs), 2) if not is_books else 0
            g_tax= round(sum(r['tax_amount']    for r in objs), 2) if not is_books else 0
            b_tax= round(sum(b['tax_amount'] for b in books_list), 2) if is_books else \
                   round(sum(r.get('books_tax') or 0 for r in objs), 2)
            diff = round(g_tax - b_tax, 2) if not is_books else 0

            row_vals = [cat, cnt, g_tx if not is_books else '', g_tax if not is_books else '', b_tax, diff if not is_books else '']
            _data_row(ws, rn, row_vals, bg=bg, amt_cols={3,4,5,6})

        r = start_row + 2
        _row(r,   'Matched',               matched,    [],    GREEN_ROW)
        _row(r+1, 'Unmatched (review)',    unmatched,  [],    YELLOW_ROW)
        _row(r+2, 'In GSTR-2B, Not Books', not_books, [],    BLUE_ROW)
        _row(r+3, 'In Books, Not GSTR-2B', [],        books, PURPLE_ROW, is_books=True)

        # Total row
        all_g_tax = round(sum(r['tax_amount'] for r in matched + unmatched + not_books), 2)
        all_b_tax = round(sum((r.get('books_tax') or 0) for r in matched + unmatched)
                         + sum(b['tax_amount'] for b in books), 2)
        diff      = round(all_g_tax - all_b_tax, 2)
        tot_row   = r + 4
        tot_vals  = ['TOTAL', len(matched)+len(unmatched)+len(not_books)+len(books), '',
                     all_g_tax, all_b_tax, diff]
        for ci, v in enumerate(tot_vals, 1):
            tc = ws.cell(row=tot_row, column=ci, value=v)
            tc.fill = _fill(NAVY); tc.font = _font(bold=True, color=WHITE_HEX)
            tc.alignment = Alignment(vertical='center')
            tc.border = _border()
            if ci in {3, 4, 5}: tc.number_format = '#,##0.00'
            if ci == 6:
                tc.number_format = '#,##0.00'
                if abs(diff) > 1:
                    tc.fill = _fill(RED_TOTAL); tc.font = _font(bold=True, color=WHITE_HEX)
        ws.row_dimensions[tot_row].height = 18
        return tot_row + 2

    next_r = _proof_table(4,  '■ CGST (Intra-State)', mc, uc, nc, bc)
    _proof_table(next_r, '■ IGST (Inter-State)',  mi, ui, ni, bi)

    _autofit(ws)
    _freeze(ws, 3, 1)

# ── Sheet 2: Matched ──────────────────────────────────────────────────────────
def build_matched_sheet(wb, entries):
    ws = wb.create_sheet("Matched")
    _title_row(ws, 1, "✅  MATCHED ENTRIES", 14)
    hdrs = ['Invoice No','Doc','Vendor Name','GSTIN','Date','Taxable ₹',
            'GSTR-2B Tax ₹','Ledger','Books Party','Books Date','Books Tax ₹','Diff ₹','Tax Type','Match Type']
    _header_row(ws, 2, hdrs)

    for i, r in enumerate(entries):
        bg = ALT_GREEN if i % 2 == 0 else WHITE_HEX
        vals = [r['invoice_no'], r.get('doc_type','Invoice'), r['supplier_name'], r['gstin'], r['invoice_date'],
                r['taxable_value'], r['tax_amount'], r.get('books_ledger',''),
                r.get('books_party',''), r.get('books_date',''),
                r.get('books_tax') or 0, r.get('tax_diff', 0),
                r['tax_type'], r.get('match_type','')]
        _data_row(ws, i + 3, vals, bg=bg, amt_cols={6,7,11,12})

    _autofit(ws); _freeze(ws, 3, 1); _autofilter(ws, 2, 14)

# ── Sheet 3: Unmatched ────────────────────────────────────────────────────────
def build_unmatched_sheet(wb, entries):
    ws = wb.create_sheet("Unmatched")
    _title_row(ws, 1, "⚠️  UNMATCHED ENTRIES — Present in both but with discrepancies", 13)
    hdrs = ['Invoice No','Doc','Vendor Name (2B)','GSTIN','Date','GSTR-2B Tax ₹',
            'Books Party','Books Ledger','Books Date','Books Tax ₹','Diff ₹','Tax Type','Discrepancy Reason']
    _header_row(ws, 2, hdrs)

    for i, r in enumerate(entries):
        bg = ALT_AMBER if i % 2 == 0 else WHITE_HEX
        diff = r.get('tax_diff', 0)
        vals = [r['invoice_no'], r.get('doc_type','Invoice'), r['supplier_name'], r['gstin'], r['invoice_date'],
                r['tax_amount'], r.get('books_party',''), r.get('books_ledger',''),
                r.get('books_date',''), r.get('books_tax') or 0,
                diff, r['tax_type'], r.get('discrepancy_reason','')]
        _data_row(ws, i + 3, vals, bg=bg, amt_cols={6,10,11})
        if diff and diff > 0:
            ws.cell(row=i+3, column=11).fill = _fill(RED_DIFF)
            ws.cell(row=i+3, column=11).font = _font(bold=True, color=RED_TOTAL)

    _autofit(ws); _freeze(ws, 3, 1); _autofilter(ws, 2, 13)

# ── Sheet 4: In GSTR-2B, Not in Books ────────────────────────────────────────
def build_not_in_books_sheet(wb, entries):
    ws = wb.create_sheet("Not In Books")
    _title_row(ws, 1, "📋  IN GSTR-2B BUT NOT IN PURCHASE REGISTER — Book these to claim ITC", 11)
    hdrs = ['Invoice No','Doc','Vendor Name','GSTIN','Date','Place of Supply',
            'Taxable ₹','Tax ₹','Tax Type','ITC Available','Action Required']
    _header_row(ws, 2, hdrs)

    for i, r in enumerate(entries):
        bg = BLUE_ROW if i % 2 == 0 else WHITE_HEX
        itc = r.get('itc_availability', 'Yes') or 'Yes'
        tax = r['tax_amount']
        mtype = r.get('match_type', '')
        if mtype in ('Offset Pair', 'Nil Tax'):
            action = r.get('discrepancy_reason', '')
        elif r.get('doc_type') == 'Credit Note':
            action = f"Credit note not booked — book to reverse ITC of ₹{tax:,.2f}"
        elif itc.lower() in ('no', 'ineligible'):
            action = "ITC NOT available — consult CA before booking"
        else:
            action = f"Book in Tally — ITC of ₹{tax:,.2f} available"
        vals = [r['invoice_no'], r.get('doc_type','Invoice'), r['supplier_name'], r['gstin'], r['invoice_date'],
                r.get('place_of_supply',''), r['taxable_value'], tax,
                r['tax_type'], itc, action]
        _data_row(ws, i + 3, vals, bg=bg, amt_cols={7,8})

    # Total row — net bookable ITC only: skip offset pairs and nil-tax rows,
    # credit notes reduce the total
    if entries:
        tr = len(entries) + 3
        total_tax = 0.0
        for r in entries:
            if r.get('match_type') in ('Offset Pair', 'Nil Tax'): continue
            if r.get('doc_type') == 'Credit Note': total_tax -= r['tax_amount']
            else:                                  total_tax += r['tax_amount']
        total_tax = round(total_tax, 2)
        for ci in range(1, 12):
            tc = ws.cell(row=tr, column=ci,
                         value=('NET BOOKABLE ITC' if ci == 1 else (total_tax if ci == 8 else '')))
            tc.fill = _fill(RED_TOTAL); tc.font = _font(bold=True, color=WHITE_HEX)
            tc.border = _border()
            if ci == 8: tc.number_format = '#,##0.00'
        ws.row_dimensions[tr].height = 18

    _autofit(ws); _freeze(ws, 3, 1); _autofilter(ws, 2, 11)

# ── Sheet 5: Others ───────────────────────────────────────────────────────────
def build_others_sheet(wb, books_cgst, books_igst, itc_no):
    ws = wb.create_sheet("Others")
    _title_row(ws, 1, "📁  OTHERS", 7)
    cur_row = 3

    def _books_section(title, note, entries):
        nonlocal cur_row
        ws.merge_cells(start_row=cur_row, start_column=1, end_row=cur_row, end_column=7)
        ws.cell(row=cur_row, column=1, value=title).fill = _fill(PURPLE_ROW)
        ws.cell(row=cur_row, column=1).font = _font(bold=True)
        ws.cell(row=cur_row, column=1).alignment = Alignment(horizontal='left')
        ws.row_dimensions[cur_row].height = 18
        cur_row += 1

        ws.merge_cells(start_row=cur_row, start_column=1, end_row=cur_row, end_column=7)
        ws.cell(row=cur_row, column=1, value=f"⚠ {note}").font = _font(bold=True, color='B45309')
        ws.row_dimensions[cur_row].height = 16
        cur_row += 1

        hdrs = ['Party Name','Narration','Date','Ledger','Books Tax ₹','Tax Type','Action']
        _header_row(ws, cur_row, hdrs); cur_row += 1

        if entries:
            for i, b in enumerate(entries):
                bg = PURPLE_ROW if i % 2 == 0 else WHITE_HEX
                vals = [b.get('party',''), b.get('narration',''), b.get('date',''),
                        b.get('ledger',''), b.get('tax_amount',0), b.get('tax_type',''),
                        'Chase supplier to file GSTR-1 — ITC blocked']
                _data_row(ws, cur_row, vals, bg=bg, amt_cols={5}); cur_row += 1
        else:
            ws.merge_cells(start_row=cur_row, start_column=1, end_row=cur_row, end_column=7)
            ws.cell(row=cur_row, column=1, value='No entries').font = _font(color='9CA3AF')
            cur_row += 1
        cur_row += 1

    _books_section('A1 — In Books, Not in GSTR-2B (CGST)',
                   'Chase supplier to file GSTR-1 — ITC blocked', books_cgst)
    _books_section('A2 — In Books, Not in GSTR-2B (IGST)',
                   'Credit notes or unmatched IGST entries — review', books_igst)

    # Section B: ITC Availability = No
    ws.merge_cells(start_row=cur_row, start_column=1, end_row=cur_row, end_column=7)
    ws.cell(row=cur_row, column=1, value='B — ITC Blocked Entries (ITC Availability = No in GSTR-2B)').fill = _fill(YELLOW_ROW)
    ws.cell(row=cur_row, column=1).font = _font(bold=True)
    ws.cell(row=cur_row, column=1).alignment = Alignment(horizontal='left')
    ws.row_dimensions[cur_row].height = 18; cur_row += 1

    hdrs_b = ['Invoice No','Vendor Name','GSTIN','Date','Tax ₹','ITC Availability','Note']
    _header_row(ws, cur_row, hdrs_b); cur_row += 1
    if itc_no:
        for i, r in enumerate(itc_no):
            bg = YELLOW_ROW if i % 2 == 0 else WHITE_HEX
            vals = [r['invoice_no'], r['supplier_name'], r['gstin'], r['invoice_date'],
                    r['tax_amount'], r.get('itc_availability','No'),
                    'ITC not available per GSTR-2B — do not claim']
            _data_row(ws, cur_row, vals, bg=bg, amt_cols={5}); cur_row += 1
    else:
        ws.merge_cells(start_row=cur_row, start_column=1, end_row=cur_row, end_column=7)
        ws.cell(row=cur_row, column=1, value='No ITC-blocked entries').font = _font(color='9CA3AF')

    _autofit(ws); _freeze(ws, 3, 1)

# ── Phase 2: generate Excel ───────────────────────────────────────────────────
def run_generate(parse_result, claude_results, output_path):
    final    = apply_claude_results(parse_result, claude_results)
    results  = final['results']
    books    = final['books_only']
    company  = final.get('company_info', {})

    mc = [r for r in results if r['status']=='matched'      and r['tax_type']=='CGST']
    uc = [r for r in results if r['status']=='unmatched'    and r['tax_type']=='CGST']
    nc = [r for r in results if r['status']=='not_in_books' and r['tax_type']=='CGST']
    mi = [r for r in results if r['status']=='matched'      and r['tax_type']=='IGST']
    ui = [r for r in results if r['status']=='unmatched'    and r['tax_type']=='IGST']
    ni = [r for r in results if r['status']=='not_in_books' and r['tax_type']=='IGST']
    bc = [b for b in books if b.get('tax_type','CGST')=='CGST']
    bi = [b for b in books if b.get('tax_type','CGST')=='IGST']
    itc_no = [r for r in results if (r.get('itc_availability') or '').upper() in ('NO','INELIGIBLE')]

    wb = openpyxl.Workbook()
    wb.remove(wb.active)

    build_summary_sheet(wb, company, mc, uc, nc, bc, mi, ui, ni, bi)
    build_matched_sheet(wb, mc + mi)
    build_unmatched_sheet(wb, uc + ui)
    build_not_in_books_sheet(wb, nc + ni)
    build_others_sheet(wb, bc, bi, itc_no)

    wb.save(output_path)

    def _s(matched, unmatched, notbook, bk):
        g_tax = round(sum(r['tax_amount'] for r in matched+unmatched+notbook), 2)
        b_tax = round(sum((r.get('books_tax') or 0) for r in matched+unmatched)
                     + sum(b['tax_amount'] for b in bk), 2)
        return {
            'matched':      len(matched),
            'unmatched':    len(unmatched),
            'not_in_books': len(notbook),
            'books_only':   len(bk),
            'gstr2b_total': g_tax,
            'books_total':  b_tax,
            'difference':   round(g_tax - b_tax, 2),
        }

    entries = [{k: v for k, v in r.items()
                if not k.startswith('_') and k not in ('tier2_5', 'rate_label')}
               for r in results]

    return {
        'status': 'ok',
        'summary': {
            'company': company.get('legal_name') or company.get('trade_name', ''),
            'gstin':   company.get('gstin', ''),
            'period':  company.get('period', ''),
            'fy':      company.get('fy', ''),
            'cgst':    _s(mc, uc, nc, bc),
            'igst':    _s(mi, ui, ni, bi),
        },
        'entries': entries,
    }

# ── Main ──────────────────────────────────────────────────────────────────────
def main():
    try:
        payload = json.loads(sys.stdin.read())
        phase   = payload.get('phase', 'parse')

        if phase == 'parse':
            result = run_parse(payload['gstr2b_path'], payload['ledger_paths'])
        elif phase == 'generate':
            result = run_generate(
                payload['parse_result'],
                payload.get('claude_results', []),
                payload['output_path'],
            )
        else:
            result = {'error': f'Unknown phase: {phase}'}

        sys.stdout.write(json.dumps(result, default=str))
        sys.stdout.flush()
    except Exception as e:
        import traceback
        sys.stdout.write(json.dumps({'error': str(e), 'traceback': traceback.format_exc()}))
        sys.stdout.flush()
        sys.exit(1)

if __name__ == '__main__':
    main()
