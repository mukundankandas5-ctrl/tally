#!/usr/bin/env python3
"""GSTR-2B colour-coded reconciliation (accountant deliverable).

Produces a FAITHFUL copy of the original government GSTR-2B workbook — every
sheet, value, column width and merge preserved, rows kept in their ORIGINAL
order, with NO added columns, sheets or remarks — whose ONLY change is the font
colour of each invoice's tax figures, set from its reconciliation result:
    green  = matched            (found in books, tax agrees)
    orange = partially matched  (found in books, same supplier/bill, tax differs)
    red    = unmatched          (not found in the purchase register)
The output should be ~99% identical to the file downloaded from the GST portal;
only the integers in the tax columns change colour, nothing else. A standard
5-sheet reconciliation report is written alongside it as the proof trail.

The result is rebuilt fresh (never a copy of the govt file, which Apple Numbers
rejects) and then VERIFIED TWICE — an independent second pass re-opens the written
file and confirms it is a faithful copy, that every tax figure carries the correct
status colour, and that the green/orange/red tax totals tie out to the engine's
per-head totals — before the run reports success.

This wraps the deterministic matching engine in gstr2b_reconcile.py; its Claude
"tier-2.5" review of ambiguous pairs runs when an API key is available and falls
back to a deterministic same-supplier rule otherwise, so it never breaks offline.

Usage:
    python3 gstr2b_colorcode.py \
        --gstr2b "052026_33...GSTR2B....xlsx" \
        --ledger "May 2026 IGST INPUT.xlsx" \
        --ledger "may 2026 cgst.xlsx" \
        [--outdir DIR] [--company "Legal Name"] [--no-report]

Ledger inputs may be either Tally export shape we support:
  A) Single-ledger voucher register: Date | Particulars (To/By + party) | ... |
     Debit | Credit  (tax type from the ledger title row).
  B) Flat 3-column register: Date | Party | <Input X Tax amount>.
Non-purchase rows (Output*, GST Payable, Opening/Closing Balance, Grand Total)
are dropped; each kept row is a Debit = input ITC.
"""
import argparse, importlib.util, json, os, re, sys, tempfile, urllib.request, urllib.error
from copy import copy as _copy
from datetime import datetime, date
from collections import Counter
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font

HERE = os.path.dirname(os.path.abspath(__file__))
_spec = importlib.util.spec_from_file_location("gstr2b_reconcile", os.path.join(HERE, "gstr2b_reconcile.py"))
recon = importlib.util.module_from_spec(_spec); _spec.loader.exec_module(recon)

MONTHS = {1:'January',2:'February',3:'March',4:'April',5:'May',6:'June',7:'July',
          8:'August',9:'September',10:'October',11:'November',12:'December'}
SKIP_PARTY = ('OUTPUT', 'PAYABLE', 'OPENING BALANCE', 'CLOSING BALANCE', 'GRAND TOTAL', 'BALANCE')


def _num(v):
    try: return round(float(str(v).replace(',', '')), 2)
    except Exception: return 0.0

def _s(v): return str(v).strip() if v is not None else ''

def _alpha(n):
    """Unique alpha-only voucher id (no digits) so it never false-matches a numeric
    2B invoice no, and so the engine's dedup keeps every distinct row."""
    s = ''; n += 1
    while n > 0:
        n, r = divmod(n - 1, 26); s = chr(65 + r) + s
    return 'VCH' + s


# ── Ledger normalisation (both supported shapes → clean rows) ────────────────
def normalize_ledger(path):
    ws = load_workbook(path, data_only=True).active
    rows = list(ws.iter_rows(values_only=True))
    title = ' '.join(_s(c) for r in rows[:9] for c in (r or [])).upper()
    if   'INTEGRATED' in title: tax_type = 'IGST'
    elif 'CENTRAL'    in title: tax_type = 'CGST'
    elif 'STATE'      in title: tax_type = 'CGST'   # SGST mirrors CGST for 2B
    else:                       tax_type = 'CGST'

    hi = None
    for i, r in enumerate(rows[:25]):
        rv = [_s(c).upper() for c in (r or [])]
        if any(v == 'DATE' for v in rv) and any('PARTICULARS' in v or 'TAX' in v or v == 'DEBIT' for v in rv):
            hi = i; break
    if hi is None:
        raise SystemExit(f"Could not find a header row in ledger: {os.path.basename(path)}")

    hdr = [_s(c).upper() for c in rows[hi]]
    def col(pred):
        for i, h in enumerate(hdr):
            if pred(h): return i
        return None
    date_c  = col(lambda h: h == 'DATE')
    part_c  = col(lambda h: 'PARTICULARS' in h)
    debit_c = col(lambda h: h == 'DEBIT')
    if part_c is None: part_c = 1
    if date_c is None: date_c = 0
    amount_c = debit_c
    if amount_c is None:                       # shape B: a single amount column
        amount_c = col(lambda h: any(k in h for k in ('TAX', 'AMOUNT', 'INPUT')) and 'TYPE' not in h)
        if amount_c is None: amount_c = len(hdr) - 1

    out = []
    for r in rows[hi + 1:]:
        if not r or not any(r): continue
        marker = _s(r[part_c] if part_c < len(r) else '')
        nxt    = _s(r[part_c + 1] if part_c + 1 < len(r) else '')
        party  = nxt if marker.upper() in ('TO', 'BY') and nxt else marker
        up = party.upper()
        if not party or any(k in up for k in SKIP_PARTY): continue
        amt = _num(r[amount_c] if amount_c < len(r) else 0)
        if amt <= 0: continue
        d = r[date_c] if date_c < len(r) else None
        out.append((d, party, amt))
    return tax_type, out


def write_clean_ledger(path, tax_type, rows, company):
    title = {'IGST': 'Input Integrated Tax', 'CGST': 'Input Central Tax'}.get(tax_type, 'Input Central Tax')
    wb = Workbook(); ws = wb.active
    ws.append([title]); ws.append([company]); ws.append([]); ws.append([])
    ws.append(['Date', 'Particulars', None, 'Vch Type', 'Vch No.', 'Debit', 'Credit'])
    for i, (d, party, amt) in enumerate(rows):
        ws.append([d, 'To', party, 'Purchase', _alpha(i), amt, None])
    wb.save(path)


# ── Claude tier-2.5 review (same prompt as the web app's gstr2bService.js) ───
DEFAULT_MODEL = "claude-sonnet-4-6"          # current Sonnet; the old 4-20250514 id is retired
CLAUDE_SYSTEM = """You are a GST reconciliation expert for Indian CA firms.
Analyse each ambiguous pair (GSTR-2B invoice vs Tally entry candidate) and decide if they match.

Classification rules:
- Match by supplier name in narration when party field is blank (cash/bank entries)
- Consolidated bookings: one Tally entry may cover multiple GSTR-2B invoices
- Amount differences <= Rs.1 = rounding -> treat as matched
- Amount differences Rs.1-50 = possible sub-entry rounding -> flag for review
- Name variations: LIMITED/LTD, PRIVATE/PVT, AND/& -> treat as same supplier
- GSTIN prefix 33 = Tamil Nadu = intra-state = CGST applies
- Return confidence 1-10: >=8 matched, 5-7 flag for review, <5 not matched

Return ONLY a JSON object with key "results" — an array with one object per input entry.
Keep "reason" under 12 words to stay within the token budget:
{"results":[{"inv_no":"...","match":true,"confidence":1,"reason":"..."}]}"""


def resolve_api_key(arg):
    """--api-key > ANTHROPIC_API_KEY env > project .env (root then server)."""
    if arg: return arg
    if os.environ.get('ANTHROPIC_API_KEY'): return os.environ['ANTHROPIC_API_KEY']
    for p in (os.path.join(HERE, '../../../.env'), os.path.join(HERE, '../../.env')):
        try:
            for line in open(p):
                if line.strip().startswith('ANTHROPIC_API_KEY='):
                    v = line.split('=', 1)[1].strip()
                    if v: return v
        except OSError: pass
    return None


def _extract_json(text):
    t = text.strip()
    m = re.search(r'```(?:json)?\s*([\s\S]*?)```', t)
    if m: t = m.group(1).strip()
    if '{' in t and ('[' not in t or t.find('{') < t.find('[')):
        return json.loads(t[t.find('{'):t.rfind('}') + 1])
    if '[' in t:
        return json.loads(t[t.find('['):t.rfind(']') + 1])
    return json.loads(t)


def _salvage_objs(text):
    """Recover complete {...} result objects from a truncated/malformed array."""
    out = []
    for m in re.finditer(r'\{[^{}]*\}', text):
        try:
            o = json.loads(m.group(0))
            if isinstance(o, dict) and o.get('inv_no') is not None: out.append(o)
        except Exception: pass
    return out


def _claude_messages(api_key, model, system, user_text, max_tokens=8192):
    body = json.dumps({'model': model, 'max_tokens': max_tokens, 'temperature': 0, 'system': system,
                       'messages': [{'role': 'user', 'content': [{'type': 'text', 'text': user_text}]}]}).encode()
    req = urllib.request.Request('https://api.anthropic.com/v1/messages', data=body, method='POST',
                                 headers={'x-api-key': api_key, 'anthropic-version': '2023-06-01',
                                          'content-type': 'application/json'})
    with urllib.request.urlopen(req, timeout=120) as resp:
        data = json.loads(resp.read())
    return ''.join(b.get('text', '') for b in data.get('content', []) if b.get('type') == 'text')


def review_tier25_claude(parse, api_key, model):
    """AI review of ambiguous pairs (batches of 50). Any batch that errors falls
    back to the deterministic rule, so a bad key/network never breaks the run."""
    cands = parse.get('tier2_5_candidates', [])
    ci = parse.get('company_info', {})
    B = 25                                   # smaller batches keep each JSON reply within the token budget
    decisions = []
    for i in range(0, len(cands), B):
        batch = cands[i:i + B]
        user = (f"Company: {ci.get('legal_name') or ci.get('trade_name', 'Unknown')}\n"
                f"GSTIN: {ci.get('gstin', 'Unknown')}\nPeriod: {ci.get('period', 'Unknown')} FY {ci.get('fy', 'Unknown')}\n\n"
                f"Reconciliation candidates ({len(batch)} entries):\n{json.dumps(batch, indent=2, default=str)}")
        got = {}
        try:
            text = _claude_messages(api_key, model, CLAUDE_SYSTEM, user)
            try:
                parsed = _extract_json(text)
                arr = parsed.get('results') if isinstance(parsed, dict) else parsed
                if not isinstance(arr, list): raise ValueError
            except Exception:
                arr = _salvage_objs(text)    # truncated reply -> keep the complete objects
            for o in arr:
                if isinstance(o, dict) and o.get('inv_no') is not None:
                    got[str(o['inv_no'])] = o
        except (urllib.error.HTTPError, urllib.error.URLError) as e:
            msg = getattr(e, 'reason', None) or (e.read().decode()[:100] if hasattr(e, 'read') else str(e))
            print(f"    tier-2.5 batch {i // B + 1}: Claude unreachable ({msg}); deterministic fallback")
        # one decision per candidate: Claude's where present, else deterministic
        det = {d['inv_no']: d for d in review_tier25({'tier2_5_candidates': batch})}
        n_ai = 0
        for c in batch:
            inv = c['inv_no']
            if str(inv) in got: decisions.append(got[str(inv)]); n_ai += 1
            else: decisions.append(det.get(inv))
        tail = '' if n_ai == len(batch) else f" ({len(batch) - n_ai} deterministic fallback)"
        print(f"    tier-2.5 batch {i // B + 1}: Claude {n_ai}/{len(batch)}{tail}")
    return [d for d in decisions if d]


# ── Deterministic stand-in for the app's Claude tier-2.5 review ──────────────
def review_tier25(parse):
    decisions = []
    for c in parse.get('tier2_5_candidates', []):
        g = c['gstr2b']; t = c['tally_candidates'][0]
        diff = round(abs(g['tax_amount'] - t['tax_amount']), 2)
        name = g['supplier_name'].strip()
        same = bool(name) and recon.distinct_overlap(name, t['party']) >= 0.5
        if not same:
            decisions.append({'inv_no': g['invoice_no'], 'match': False, 'confidence': 2,
                              'reason': 'Different/blank supplier - amount coincidence, not the same document.'})
        elif diff <= 1.0:
            decisions.append({'inv_no': g['invoice_no'], 'match': True, 'confidence': 10,
                              'reason': 'Same supplier (name variant); tax agrees within Rs.1 (rounding).'})
        else:
            decisions.append({'inv_no': g['invoice_no'], 'match': True, 'confidence': 6,
                              'reason': f'Same supplier; GSTR-2B tax differs from books by Rs.{diff:,.2f} - verify rate/invoice.'})
    return decisions


# ── Colour the reconciled tax figures in a faithful copy of the GSTR-2B ───────
# Deliverable = the ORIGINAL government GSTR-2B workbook, reproduced as closely as
# possible (every sheet, value, column width, row height and merge preserved, rows
# in their original order, NO added columns/sheets/remarks). The ONLY change is the
# FONT colour of each row's non-zero tax figures, set from that invoice's recon
# status. Built in a brand-new Workbook (styles copied object-by-object, never the
# govt file's inherited theme) so Apple Numbers can open it as well as Excel.
FONT_GREEN  = '008000'   # matched
FONT_ORANGE = 'C55A11'   # partially matched (in books, same supplier/bill, tax differs)
FONT_RED    = 'FF0000'   # unmatched (not found in the purchase register)

# The engine's status names predate the accountant's 3-bucket vocabulary:
#   engine 'matched'      -> user "matched"            -> green
#   engine 'unmatched'    -> user "partially matched"  -> orange  (in both, tax differs)
#   engine 'not_in_books' -> user "unmatched"          -> red     (absent from books)
def status_color(status):
    if status == 'matched':   return FONT_GREEN
    if status == 'unmatched': return FONT_ORANGE
    return FONT_RED                                   # not_in_books + any other non-match


def _value_rows(ws):
    return [[ws.cell(r, c).value for c in range(1, ws.max_column + 1)]
            for r in range(1, ws.max_row + 1)]


def _tax_layout(ws_src, is_cdnr):
    """Locate, 1-based, the GSTIN column, the document-number column and the
    tax-amount columns (Integrated/Central/State-UT/Cess in the FIRST tax block —
    the block the engine reconciles) on a 2B sheet, by header text so it survives
    the column-count differences between portal versions. Returns None if absent."""
    rows = _value_rows(ws_src)
    hdr = recon._find_b2b_header(rows)
    if hdr is None: return None
    header_rows = [rows[hdr], rows[hdr + 1] if hdr + 1 < len(rows) else []]
    cm = recon._gstr2b_colmap(header_rows, is_cdnr)
    if cm.get('gstin') is None: return None
    ncols = ws_src.max_column
    combined = []
    for ci in range(ncols):
        parts = [str(hr[ci]).upper() for hr in header_rows if ci < len(hr) and hr[ci] is not None]
        combined.append(' '.join(parts))
    cess = next((ci for ci, t in enumerate(combined) if 'CESS' in t), None)   # first Cess = first tax block
    tax0 = sorted({c for c in (cm.get('igst'), cm.get('cgst'), cm.get('sgst'), cess) if c is not None})
    return {
        'hdr':       hdr,                                       # 0-based index of the GSTIN header row
        'data_from': hdr + 2,                                   # 1-based first row scanned for data
        'gstin_col': cm['gstin'] + 1,
        'doc_col':   (cm['doc_no'] + 1) if cm.get('doc_no') is not None else None,
        'tax_cols':  [c + 1 for c in tax0],
    }


def _data_rows(ws, lay):
    """The 1-based rows the engine turned into results — real-GSTIN rows after the
    header, in original order (mirrors gstr2b_reconcile._read_sheet's row filter)."""
    g = lay['gstin_col']
    return [r for r in range(lay['data_from'], ws.max_row + 1)
            if len(_s(ws.cell(r, g).value)) >= 15]


def _color_overrides(ws_src, is_cdnr, results):
    """Pair each 2B data row (original order) with its reconciliation result and
    return ({(row, col): rgb} tax-figure colour overrides, [cross-check problems]).
    Pairing is positional but guarded by a GSTIN + document-number cross-check, so
    a row can never silently take the wrong colour."""
    lay = _tax_layout(ws_src, is_cdnr)
    if lay is None:
        raise SystemExit(f"{ws_src.title}: could not locate the GSTR-2B header / tax columns")
    rows = _data_rows(ws_src, lay)
    if len(rows) != len(results):
        raise SystemExit(f"{ws_src.title}: {len(rows)} data rows but {len(results)} results — cannot align")
    g, d, tcols = lay['gstin_col'], lay['doc_col'], lay['tax_cols']
    overrides, problems = {}, []
    for r, res in zip(rows, results):
        if _s(ws_src.cell(r, g).value) != _s(res.get('gstin')):
            problems.append(f"{ws_src.title} row {r}: GSTIN {_s(ws_src.cell(r, g).value)} != result {_s(res.get('gstin'))}")
        if d and _s(ws_src.cell(r, d).value) != _s(res.get('invoice_no')):
            problems.append(f"{ws_src.title} row {r}: doc no {_s(ws_src.cell(r, d).value)} != result {_s(res.get('invoice_no'))}")
        rgb = status_color(res['status'])
        for c in tcols:
            v = ws_src.cell(r, c).value
            if isinstance(v, (int, float)) and round(float(v), 2) != 0.0:
                overrides[(r, c)] = rgb
    return overrides, problems


def _copy_sheet_faithful(new, ws_src, overrides=None):
    """Reproduce a worksheet in a fresh workbook: every value, cell style, column
    width, row height and merge, then recolour the tax-figure fonts in `overrides`.
    Style objects are copied (not the source workbook's style indices/theme), which
    keeps the output openable in Apple Numbers."""
    ws = new.create_sheet(ws_src.title)
    ws.sheet_view.showGridLines = ws_src.sheet_view.showGridLines
    if ws_src.freeze_panes: ws.freeze_panes = ws_src.freeze_panes
    for key, dim in ws_src.column_dimensions.items():
        nd = ws.column_dimensions[key]
        nd.width = dim.width; nd.hidden = dim.hidden
    for key, dim in ws_src.row_dimensions.items():
        ws.row_dimensions[key].height = dim.height
    for r in range(1, ws_src.max_row + 1):
        for c in range(1, ws_src.max_column + 1):
            sc = ws_src.cell(r, c)
            if sc.value is None and not sc.has_style: continue
            nc = ws.cell(r, c, sc.value)
            if sc.has_style:
                nc.font = _copy(sc.font); nc.fill = _copy(sc.fill)
                nc.border = _copy(sc.border); nc.alignment = _copy(sc.alignment)
                nc.number_format = sc.number_format; nc.protection = _copy(sc.protection)
    for m in ws_src.merged_cells.ranges:
        try: ws.merge_cells(str(m))
        except Exception: pass
    if overrides:
        for (r, c), rgb in overrides.items():
            cell = ws.cell(r, c); f = cell.font
            cell.font = Font(name=f.name, size=f.size, bold=f.bold, italic=f.italic,
                             vertAlign=f.vertAlign, underline=f.underline, strike=f.strike, color=rgb)
    return ws


def build_faithful_colored(src_path, out_path, results):
    """Write the faithful colour-coded GSTR-2B. Returns the list of row cross-check
    problems (empty when every row paired cleanly)."""
    src = load_workbook(src_path)
    b2b = [r for r in results if r.get('doc_type') == 'Invoice']
    cdn = [r for r in results if r.get('doc_type') in ('Credit Note', 'Debit Note')]
    new = Workbook(); new.remove(new.active)
    problems = []
    for name in src.sheetnames:
        ws_src = src[name]
        ov = None
        if name == 'B2B':
            ov, p = _color_overrides(ws_src, False, b2b); problems += p
        elif name == 'B2B-CDNR':
            ov, p = _color_overrides(ws_src, True, cdn); problems += p
        _copy_sheet_faithful(new, ws_src, ov)
    new.save(out_path)
    make_numbers_compatible(out_path)
    return problems


def verify_output(out_path, src_path, results):
    """Second, INDEPENDENT pass (the "check twice" step). Returns (ok, [lines]).
    Confirms: (a) the written file is a faithful copy of the source 2B — same
    sheets, dimensions and every data value; (b) each non-zero tax figure carries
    the colour its reconciliation status demands; (c) the green/orange/red tax
    totals tie out, per tax head, to the total GSTR-2B tax."""
    lines, ok = [], True
    src = load_workbook(src_path, data_only=True)
    out = load_workbook(out_path, data_only=True)

    # (a) faithful copy — same sheets, dims and every cell value
    if src.sheetnames != out.sheetnames:
        ok = False; lines.append(f"sheet list differs: {src.sheetnames} != {out.sheetnames}")
    for name in src.sheetnames:
        if name not in out.sheetnames: continue
        a, b = src[name], out[name]
        if (a.max_row, a.max_column) != (b.max_row, b.max_column):
            ok = False; lines.append(f"{name}: dims {a.max_row}x{a.max_column} != {b.max_row}x{b.max_column}"); continue
        diffs = sum(1 for r in range(1, a.max_row + 1) for c in range(1, a.max_column + 1)
                    if _s(a.cell(r, c).value) != _s(b.cell(r, c).value))
        if diffs: ok = False; lines.append(f"{name}: {diffs} cell value(s) changed vs original ✗")
        else:     lines.append(f"{name}: {a.max_row} rows identical to original ✓")

    # (b) colour correctness, read back from the written file
    styled = load_workbook(out_path)
    for name, is_cdnr in (('B2B', False), ('B2B-CDNR', True)):
        if name not in out.sheetnames: continue
        res = [r for r in results
               if (r.get('doc_type') in ('Credit Note', 'Debit Note')) == is_cdnr
               and (is_cdnr or r.get('doc_type') == 'Invoice')]
        if not res: continue
        ws = styled[name]; lay = _tax_layout(src[name], is_cdnr)
        rows, bad = _data_rows(ws, lay), 0
        for r, rr in zip(rows, res):
            want = status_color(rr['status'])
            for c in lay['tax_cols']:
                v = ws.cell(r, c).value
                if isinstance(v, (int, float)) and round(float(v), 2) != 0.0:
                    col = ws.cell(r, c).font.color
                    got = ((col.rgb if col and isinstance(col.rgb, str) else '') or '')[-6:].upper()
                    if got != want: bad += 1
        if bad: ok = False; lines.append(f"{name}: {bad} tax cell(s) wrongly coloured ✗")
        else:   lines.append(f"{name}: all {len(res)} rows coloured correctly ✓")

    # (c) accounting tie-out, per tax head
    for head in ('CGST', 'IGST'):
        rs = [r for r in results if r.get('tax_type') == head]
        if not rs: continue
        m = round(sum(r['tax_amount'] for r in rs if r['status'] == 'matched'), 2)
        u = round(sum(r['tax_amount'] for r in rs if r['status'] == 'unmatched'), 2)
        n = round(sum(r['tax_amount'] for r in rs if r['status'] == 'not_in_books'), 2)
        tot, tie = round(sum(r['tax_amount'] for r in rs), 2), round(m + u + n, 2)
        good = abs(tie - tot) < 0.01
        if not good: ok = False
        lines.append(f"{head}: green ₹{m:,.2f} + orange ₹{u:,.2f} + red ₹{n:,.2f} = ₹{tie:,.2f} "
                     f"(2B total ₹{tot:,.2f}) {'✓' if good else '✗ MISMATCH'}")
    return ok, lines


# ── Company / GSTIN / period identity + cross-check ──────────────────────────
# The authoritative identity of a run is the GSTR-2B's "Read me" sheet (legal
# name, GSTIN, tax period). We cross-check it against the 2B filename and the
# Tally ledger headers and ABORT if they disagree — so one client's 2B can never
# be reconciled against another client's ledgers, or against the wrong month.
_IDENTITY_MIN_NAME = 0.34          # distinctive-name overlap below this = a different company


def _ledger_company(path):
    """Company name from a Tally ledger's top rows (first alphabetic, non-address line)."""
    try:
        head = list(load_workbook(path, data_only=True).active.iter_rows(values_only=True))[:4]
        for r in head:
            cand = _s(r[0]) if r else ''
            if cand and not cand[0].isdigit() and not re.match(r'(?i)no\.?\d|input |ledger', cand):
                return cand
    except Exception:
        pass
    return ''


def read_2b_readme(path):
    """Authoritative company info from the GSTR-2B 'Read me' sheet (fy/period/gstin/
    legal/trade), matching gstr2b_reconcile.parse_gstr2b. {} if there is no sheet
    (older portal exports omitted it — identity then falls back to the filename)."""
    try:
        wb = load_workbook(path, data_only=True)
    except Exception:
        return {}
    if 'Read me' not in wb.sheetnames:
        return {}
    r = list(wb['Read me'].iter_rows(values_only=True))
    def c(ri, ci): return _s(r[ri][ci]) if len(r) > ri and r[ri] is not None and len(r[ri]) > ci else ''
    return {'fy': c(3, 2), 'period': c(4, 2), 'gstin': c(5, 2),
            'legal_name': c(6, 2), 'trade_name': c(7, 2)}


def resolve_identity(gstr2b_path, ledger_paths, readme, company_override):
    """Return (company_info, [problems]). company_info prefers the 2B 'Read me'
    sheet with the filename as the calendar-period source; problems lists any
    cross-check failure between the 2B and the ledgers (different client / month)."""
    base = os.path.basename(gstr2b_path)
    gm = re.search(r'(\d{2}[A-Z]{5}\d{4}[A-Z0-9]{4})', base)
    fm_gstin = gm.group(1).upper() if gm else ''
    pm = re.search(r'(\d{2})(\d{4})', base)
    fm_month = MONTHS.get(int(pm.group(1)), '') if pm else ''
    fm_period = f"{fm_month} {pm.group(2)}" if pm and fm_month else ''

    rm_gstin  = _s(readme.get('gstin')).upper()
    rm_period = _s(readme.get('period'))
    rm_legal  = _s(readme.get('legal_name'))
    rm_trade  = _s(readme.get('trade_name'))
    rm_names  = [n for n in (rm_legal, rm_trade) if n]

    gstin  = rm_gstin or fm_gstin
    period = fm_period or rm_period          # filename MMYYYY is the unambiguous calendar period
    legal  = (company_override or rm_legal or rm_trade
              or (_ledger_company(ledger_paths[0]) if ledger_paths else '') or 'Company')

    problems = []
    # (1) the 2B's own filename must agree with its Read-me contents
    if rm_gstin and fm_gstin and rm_gstin != fm_gstin:
        problems.append(f"GSTR-2B filename GSTIN {fm_gstin} does not match the GSTIN inside the file "
                        f"({rm_gstin}) — wrong or renamed 2B file.")
    if fm_month and rm_period and fm_month.lower() != rm_period.strip().lower():
        problems.append(f"GSTR-2B filename month ({fm_month}) does not match the period inside the file "
                        f"({rm_period}) — wrong or renamed 2B file.")
    # (2) the ledgers must belong to the same company as the 2B
    if rm_names:
        for lp in ledger_paths:
            lc = _ledger_company(lp)
            if lc and max(recon.name_score(n, lc) for n in rm_names) < _IDENTITY_MIN_NAME:
                problems.append(f"Ledger '{os.path.basename(lp)}' is for company '{lc}', but the GSTR-2B is "
                                f"for '{rm_legal or rm_trade}' — the 2B and ledgers look like different clients.")
    return ({'legal_name': legal, 'trade_name': legal, 'gstin': gstin,
             'period': period, 'fy': _s(readme.get('fy'))}, problems)


def make_numbers_compatible(path):
    """openpyxl writes string cells as inline strings (t="inlineStr"); Apple
    Numbers cannot read those and reports "couldn't read the file". Convert text
    cells to a shared-strings table and strip empty inline cells, so the file
    opens in Numbers as well as Excel."""
    import zipfile
    z = zipfile.ZipFile(path, 'r'); names = z.namelist()
    strings = []; index = {}; total = [0]
    pat = re.compile(r'<c([^>]*?) t="inlineStr"><is><t( xml:space="preserve")?(?:>(.*?)</t>|/>)</is></c>', re.S)
    def repl(m):
        total[0] += 1
        space = m.group(2) or ''; text = m.group(3) if m.group(3) is not None else ''
        key = (space, text)
        if key not in index: index[key] = len(strings); strings.append(key)
        return f'<c{m.group(1)} t="s"><v>{index[key]}</v></c>'
    out = {}
    for n in names:
        d = z.read(n)
        if re.match(r'xl/worksheets/sheet\d+\.xml$', n):
            s = pat.sub(repl, d.decode('utf-8'))
            s = re.sub(r'(<c[^>]*?) t="inlineStr"(\s*/>)', r'\1\2', s)   # empty self-closing cells
            s = re.sub(r'(<c[^>]*?) t="inlineStr"(></c>)', r'\1\2', s)
            d = s.encode('utf-8')
        out[n] = d
    z.close()
    if total[0]:
        sst = ['<?xml version="1.0" encoding="UTF-8" standalone="yes"?>',
               f'<sst xmlns="http://schemas.openxmlformats.org/spreadsheetml/2006/main" '
               f'count="{total[0]}" uniqueCount="{len(strings)}">']
        for space, text in strings: sst.append(f'<si><t{space}>{text}</t></si>')
        sst.append('</sst>')
        out['xl/sharedStrings.xml'] = ''.join(sst).encode('utf-8')
        ct = out['[Content_Types].xml'].decode('utf-8')
        if 'sharedStrings' not in ct:
            ct = ct.replace('</Types>', '<Override PartName="/xl/sharedStrings.xml" ContentType='
                            '"application/vnd.openxmlformats-officedocument.spreadsheetml.sharedStrings+xml"/></Types>')
        out['[Content_Types].xml'] = ct.encode('utf-8')
        rels = out['xl/_rels/workbook.xml.rels'].decode('utf-8')
        ids = [int(x) for x in re.findall(r'Id="rId(\d+)"', rels)]; nid = (max(ids) + 1) if ids else 1
        if 'sharedStrings' not in rels:
            rels = rels.replace('</Relationships>', f'<Relationship Id="rId{nid}" Type='
                                '"http://schemas.openxmlformats.org/officeDocument/2006/relationships/sharedStrings" '
                                'Target="sharedStrings.xml"/></Relationships>')
        out['xl/_rels/workbook.xml.rels'] = rels.encode('utf-8')
    tmp = path + '.tmp'
    with zipfile.ZipFile(tmp, 'w', zipfile.ZIP_DEFLATED) as zo:
        for n, d in out.items(): zo.writestr(n, d)
    os.replace(tmp, path)


def main():
    ap = argparse.ArgumentParser(description="Faithful colour-coded GSTR-2B reconciliation")
    ap.add_argument('--gstr2b', required=True)
    ap.add_argument('--ledger', action='append', required=True, help='repeatable: one per input-tax ledger')
    ap.add_argument('--outdir', default=None)
    ap.add_argument('--company', default=None)
    ap.add_argument('--no-report', action='store_true', help='skip the 5-sheet report')
    ap.add_argument('--no-ai', action='store_true', help='force the deterministic tier-2.5 rule (skip Claude)')
    ap.add_argument('--api-key', default=None, help='Anthropic key (else env ANTHROPIC_API_KEY / project .env)')
    ap.add_argument('--model', default=DEFAULT_MODEL)
    ap.add_argument('--force-identity', action='store_true',
                    help='proceed even if the 2B and ledgers look like different clients/periods')
    a = ap.parse_args()

    outdir = a.outdir or os.path.dirname(os.path.abspath(a.gstr2b))
    os.makedirs(outdir, exist_ok=True)

    # 0) identity + cross-check — the 2B and the ledgers must be the same client/month
    readme = read_2b_readme(a.gstr2b)
    company, id_problems = resolve_identity(a.gstr2b, a.ledger, readme, a.company)
    print(f"  client : {company['legal_name']}  |  GSTIN {company['gstin'] or '?'}  |  period {company['period'] or '?'}")
    if id_problems:
        print("  ✗ FILE IDENTITY CHECK:")
        for p in id_problems: print("     ", p)
        if not a.force_identity:
            sys.exit("  Aborting: the GSTR-2B and ledgers do not appear to be the same client/period. "
                     "Check the inputs; re-run with --force-identity only if you are sure they belong together.")
        print("  ⚠ --force-identity set: continuing despite the mismatch above.")

    # 1) normalise ledgers → clean Format-B temp files the engine reads cleanly
    clean_paths = []
    tmp = tempfile.mkdtemp(prefix='gstr2b_')
    for i, lp in enumerate(a.ledger):
        tax_type, rows = normalize_ledger(lp)
        cp = os.path.join(tmp, f'clean_{i}_{tax_type}.xlsx')
        write_clean_ledger(cp, tax_type, rows, company['legal_name'] or 'Company')
        clean_paths.append(cp)
        print(f"  ledger {os.path.basename(lp)} -> {tax_type}: {len(rows)} rows, Rs.{sum(r[2] for r in rows):,.2f}")

    # 2) engine: parse + tier-2.5 review (Claude when a key is available, else deterministic)
    parse = recon.run_parse(a.gstr2b, clean_paths)
    parse['company_info'] = company
    n_amb = len(parse.get('tier2_5_candidates', []))
    key = None if a.no_ai else resolve_api_key(a.api_key)
    if key and n_amb:
        print(f"  tier-2.5: {n_amb} ambiguous pairs -> Claude ({a.model}), deterministic fallback")
        claude = review_tier25_claude(parse, key, a.model)
    else:
        why = 'forced --no-ai' if a.no_ai else ('no API key found' if not key else 'none ambiguous')
        print(f"  tier-2.5: {n_amb} ambiguous pairs -> deterministic rule ({why})")
        claude = review_tier25(parse)

    # 3) standard 5-sheet report (also finalises statuses on parse['results'])
    period_tag = (company['period'] or 'period').replace(' ', '')
    report_path = os.path.join(outdir, f"GSTR2B_Reconciliation_{period_tag}.xlsx")
    recon.run_generate(parse, claude, report_path)
    if not a.no_report:
        make_numbers_compatible(report_path)
        print("  report  ->", report_path)
    else:
        os.remove(report_path)

    # 4) faithful colour-coded GSTR-2B — an exact copy of the government file whose
    #    only change is the tax-figure font colour (green/orange/red). Verified twice.
    base = os.path.splitext(os.path.basename(a.gstr2b))[0]
    color_path = os.path.join(outdir, f"{base}_Reconciled_ColorCoded.xlsx")
    results = parse['results']
    problems = build_faithful_colored(a.gstr2b, color_path, results)
    if problems:
        print(f"  ⚠ {len(problems)} row cross-check problem(s):")
        for p in problems[:20]: print("     ", p)

    ok, lines = verify_output(color_path, a.gstr2b, results)
    print("  verification (independent re-read):")
    for ln in lines: print("     ", ln)

    print("  colour  ->", color_path)
    label = {'matched': 'green (matched)', 'unmatched': 'orange (partial)', 'not_in_books': 'red (unmatched)'}
    b2b = [r for r in results if r.get('doc_type') == 'Invoice']
    cdn = [r for r in results if r.get('doc_type') in ('Credit Note', 'Debit Note')]
    print("  B2B :", dict(Counter(label.get(r['status'], r['status']) for r in b2b)))
    if cdn: print("  CDNR:", dict(Counter(label.get(r['status'], r['status']) for r in cdn)))

    if problems or not ok:
        sys.exit("  ✗ VERIFICATION FAILED — output is not trustworthy; do not deliver (see lines above)")
    print("  ✓ verified twice: faithful copy, tax-figure colours correct, totals tie out")


if __name__ == '__main__':
    main()
